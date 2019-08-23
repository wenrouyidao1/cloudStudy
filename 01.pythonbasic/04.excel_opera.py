
import openpyxl

# 创建新表格
workbook = openpyxl.Workbook()
workbook.save('./exclFile/py1901.xlsx')


# 打开已经存在的表格
workbook = openpyxl.load_workbook('./exclFile/py1901.xlsx')
sheet = workbook['Sheet']
print(sheet['A1'].value)

workbook.save('./exclFile/py1901.xlsx')

# 写数据入表格
file = openpyxl.load_workbook('./exclFile/py1901.xlsx')

# # -----------------------1. 第一种方式-------------------------
# sheet = file['nginx']
# sheet['B2'] = 'this nginx'
#
# # -----------------------2. 第二种方式（不建议使用）-----------------------
# sheet = file.active
# sheet['C3'] = 'this active'

sheet = file['ansible']
sheet.append((1, '2', 3, '4'))  # append里填可迭代对象

# 最后都要保存表格
file.save('./exclFile/py1901.xlsx')



